# Grounding: transcribed
# Source: Alignment_QC_preprocessing.ipynb (author repo AkilaRanjith/Molecular-Signatures-of-Resilience...); code cells only (outputs stripped).

# ---- cell 1 ----
# This Jupyter Notebook (.ipynb file) contains code for the following tasks:
#1. Construct Kallisto index
#2. Run kbcount
#3. Perform Quality control for each sample
#4. Manage metadata
#5. Detect doublets
#6. Perform harmony integration
#7. Cluster neuronal and non-neuronal subtypes

# ---- cell 2 ----
module load kallisto
module load kb
### specify input sample 
declare -a StringArray=("input samples")

for val in ${StringArray[@]}; do
    # create directory for each sample
    mkdir=path/count/$val
    cd  path/$val
    # count the abundance for each sample 
    kb count --h5ad -i path/ref_index.idx -g path/t2g.txt -x 10xv2 -o path/$val -c1 path/spliced.txt -c2 path/unspliced.txt --workflow=lamanno --filter bustools $(find -type f -name '*.fastq.gz'| sort)
    done


# ---- cell 3 ----
import numpy as np
import pandas as pd
import scanpy as sc
import anndata 
from scipy.io import mmread
from scipy.sparse import csr_matrix



list1=["input samples"]

# output file names

list2=["output filenames"]

# Read metadata from csv files
metadata=pd.read_csv("path/sample_metadata.csv")

# Add metadata information

for i in range(len(list1)):

    adata = anndata.read_h5ad("path/"+str(list1[i])+"/counts_unfiltered/adata.h5ad")
    adata.var_names_make_unique()
    S1=metadata.loc[i]
    
  # write metadata 
    list1[i]=adata
    list1[i].obs['Unique.identifier']=S1[0]
    list1[i].obs['Assay']=S1[1]
    list1[i].obs['Sequencing.platform']=S1[2]
    list1[i].obs['Read2.length']=S1[3]
    list1[i].obs['Tissue.processing.date']=S1[4]
    list1[i].obs['Donor']=S1[5]
    list1[i].obs['subject']=S1[6]
    list1[i].obs['APOE']=S1[7]
    list1[i].obs['Brain.Region']=S1[8]
    list1[i].obs['SORT']=S1[9]
    list1[i].obs['NP.Diagnosis']=S1[10]
    list1[i].obs['Braak.stage']=S1[11]
    list1[i].obs['Disease.Group']=S1[12]
    list1[i].obs['Amyloid']=S1[13]
    list1[i].obs['Brain.weight']=S1[14]
    list1[i].obs['PMI.hr.']=S1[15]
    list1[i].obs['Race']=S1[16]
    list1[i].obs['Age']=S1[17]
    list1[i].obs['Sex']=S1[18]
    list1[i].obs['RIN']=S1[19]
    
    
    
 #write h5ad file

    list1[i].write("path/"+str(list2[i])+".h5ad")

# ---- cell 4 ----
import scipy.sparse as sparse
import scipy.io as sio
import scipy.stats as stats
import numpy as np
import anndata

list1=["input sample files"]


for i in range(len(list1)):
    #load h5ad file
    adata = anndata.read_h5ad("path/"+str(list1[i])+".h5ad")
    # convert sparse matrix to mtx format
    adata.X=adata.layers["spliced"]+adata.layers["unspliced"]
    data=adata.X
    sio.mmwrite("path/"+str(list1[i])+".mtx",data)

# ---- cell 5 ----
## load R Packages
library(DropletUtils)
library(Matrix)
library(ggplot2)
library(dplyr)
library(tidyverse)
library(ggExtra)

list1<-c("input sample file names")

list_count <- length(list1)

print(list_count)

for (i in list_count)
{
    ## load matrix for each sample
    
  m <- readMM(paste0("path/",list1[i],".mtx"))
  m <- Matrix::t(m)
  m <- as(m, "dgCMatrix")
    
    ## calculate empty droplets based on barcodes 
    
  tot_counts <- Matrix::colSums(m)  
  bc_rank <- barcodeRanks(m)
  set.seed(200)
  e.out <- emptyDrops(m)
    
  ## test each cell for ambient RNA profile
  is.cell <- e.out$FDR <= 0.05
  sum(is.cell, na.rm=TRUE)
  # write e.out
    
  write.csv(bc_rank,paste("path/",list1[i],".csv"))

  write.csv(e.out,paste("path/",list1[i],".csv"))
  mypath <- file.path("path/",paste(list1[i], ".jpg", sep = ""))
    
  # log probaility plot empty droplet testing
  jpeg(file=mypath)
  par(mar=c(5,4,1,1), mfrow=c(1,2), bty="n")
    plot(e.out$Total, -e.out$LogProb, col=ifelse(is.cell, "red", "black"),
    xlab="Total UMI count", ylab="-Log Probability", cex=0.2)
    abline(v = bc_rank@metadata$inflection, col="darkgreen")
    abline(v = bc_rank@metadata$knee, col="dodgerblue")
    legend("bottomright", legend=c("Inflection", "Knee"), bty="n", 
    col=c("darkgreen", "dodgerblue"), lty=1, cex=1.2)
    
    # zoom in low UMI 
    # log probaility plot 
  
    plot(e.out$Total, -e.out$LogProb, col=ifelse(is.cell, "red", "black"),
    xlab="Total UMI count", ylab="-Log Probability", cex=0.2, xlim=c(0,2200), ylim=c(0,2200))
    abline(v = bc_rank@metadata$inflection, col="darkgreen")
    abline(v = bc_rank@metadata$knee, col="dodgerblue")
    dev.off()
    
    # cell_calling plots   
    set.seed(100)
    #set inflection point
    limit <- bc_rank@metadata$inflection  
    all.out <- emptyDrops(m, lower=limit, test.ambient=TRUE)
    
    # write output
    write.csv(all.out,paste("path/",list1[i],".csv"))
    mypath1 <- file.path("path/",paste(list1[i], ".jpg", sep = ""))
    jpeg(file=mypath1)
    
    # histogram null hypotheis
    hist(all.out$PValue[all.out$Total <= limit & all.out$Total > 0],
    xlab="P-value", main="", col="grey80")
    dev.off()
    
    
  }
  

# ---- cell 6 ----
import numpy as np
import pandas as pd
import scanpy as sc
import anndata 
from scipy.io import mmread
import scrublet as scr
import doubletdetection
from sklearn.decomposition import TruncatedSVD
import matplotlib
import matplotlib.pyplot as plt
from scipy import sparse, io
from scipy.sparse import csr_matrix


list1=["input file names"]

for i in range(len(list1)):
    
    ## read h5ad file

    adata = anndata.read_h5ad("path/"+str(list1[i])+".h5ad")
   
    ## read bc_rank,FDR
    
    bc_rank=pd.read_csv("path/"+" "+str(list1[i])+" "+".csv")
    FDR=pd.read_csv("path/"+" "+str(list1[i])+".csv")
    
    ## copy information
    
    adata.obs.index = bc_rank.index
    adata.obs.index = FDR.index
    adata.obs['rank'] = bc_rank['rank']
    adata.obs["droplet_FDR"]=FDR["FDR"]
    adata.obs["total_counts"]=bc_rank["total"]
    
    ## write the metadata 
    adata.write("path/"+str(list1[i])+".h5ad")
    
       
    
    ### read the h5ad file
    adata = anndata.read_h5ad("path/"+str(list1[i])+".h5ad")
    
   
    ## remove cells based on FDR <0.05 
    adata=adata[adata.obs['droplet_FDR'] < 0.05,:]
    
    ## merge the spliced and unspliced counts and store in nuclear layer
    
    adata.layers["nuclear"]=adata.layers["spliced"]+adata.layers["unspliced"]
    
    ### Calculate QC metrices
    adata.X = csr_matrix(adata.layers['nuclear'])
    
      # mitochondrial genes
    adata.var['mt'] = adata.var_names.str.startswith('MT-') 
    
    # ribosomal genes
    adata.var['ribo'] = adata.var_names.str.startswith(("RPS","RPL"))
    
    # hemoglobin genes.
    adata.var['hb'] = adata.var_names.str.contains(("^HB[^(P)]"))

    sc.pp.calculate_qc_metrics(adata, qc_vars=['mt','ribo','hb'],percent_top=None, log1p=False, inplace=True)
    
    mito_genes = adata.var_names.str.startswith('MT-')
    
    # For each cell compute fraction of counts in mito genes vs. all genes
    
    adata.obs['pct_counts_mt'] = np.sum(
        adata[:, mito_genes].X, axis=1).A1/np.sum(adata.X, axis=1).A1
    
    # add the total counts per cell as observations-annotation to adata
    adata.obs['n_counts'] = adata.X.sum(axis=1).A1

    ##  save the plots
    sc.pl.violin(adata, ['n_genes_by_counts', 'n_counts', 'pct_counts_mt'],jitter=0.4, multi_panel=True)
    plt.savefig("path/"+str(list1[i])+".png")
    
  

# ---- cell 7 ----
#predict doublets
    
    clf = doubletdetection.BoostClassifier(
    n_iters=10,
    clustering_algorithm="louvain",
    standard_scaling=True,
    pseudocount=0.1,
    n_jobs=-1,)
    doublets = clf.fit(adata.X).predict(p_thresh=1e-16, voter_thresh=0.5)
    doublet_score = clf.doublet_score()
    adata.obs["doublefinder"] = doublets
    adata.obs["doubletfinder_score"] = doublet_score
    ### write output
    adata.write("path/"+str(list1[i])+".h5ad")  
    
    #convergence plot
    f = doubletdetection.plot.convergence(clf, save="path/"+str(list1[i])+".pdf", show=True, p_thresh=1e-16, voter_thresh=0.5)
    f3 = doubletdetection.plot.threshold(clf, save="path/"+str(list1[i])+".pdf", show=True, p_step=6)
   


# ---- cell 8 ----
import numpy as np
import pandas as pd
import scanpy as sc
import anndata 
from scipy.io import mmread

filenames=["input here"]
adatas = [anndata.read_h5ad("path/"+filename+".h5ad") for filename in filenames]

adata = adatas[0].concatenate(adatas[1:], batch_key='filename', batch_categories=filenames)

adata.write("path/After_QC_nuclei_combined.h5ad")


# ---- cell 9 ----
import scanpy.external as sce
import anndata
import pandas as pd
import scanpy as sc
import numpy as np
from scipy.io import mmread
from sklearn.decomposition import TruncatedSVD
import matplotlib
import matplotlib.pyplot as plt
import seaborn as sns
import csv
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

adata=anndata.read_h5ad("path/After_QC_nuclei.h5ad")

#### copy the gene name instead of ENSG id
adata.var["index1"]=adata.var.index
adata.var["gene_name"]=adata.var["gene_name"].astype('object')
adata.var.loc[adata.var['gene_name'] =="",'gene_name']=adata.var['index1']
adata.var.set_index('gene_name',inplace=True)

#### Filter minimum genes
sc.pp.filter_cells(adata, min_genes=300)
print("After removing min_genes",adata.n_obs, adata.n_vars)


#### scale and normalize data
sc.pp.normalize_total(adata, target_sum=1e4)
sc.pp.log1p(adata)

### Find HVG
sc.pp.highly_variable_genes(adata, min_mean=0.0125, max_mean=3, min_disp=0.5)

#### PCA 
sc.tl.pca(adata, svd_solver='arpack')
sc.pl.pca_variance_ratio(adata, log=True)

### harmony integration

sce.pp.harmony_integrate(adata, ['Assay','Brain.Region'])

#### check the Elbow plot for PCA
### generate leiden graph based KNN

sc.pp.neighbors(adata, use_rep="X_pca_harmony",n_pcs=35)
sc.tl.leiden(adata,resolution = 1.0, key_added = "leiden_1.0")
sc.tl.umap(adata)
sc.pl.umap(adata,color=["Assay", "leiden_1.0"],ncols=2, frameon=False,)
plt.savefig("path/After_Assay_brainregion.png")

### List the major celltype marker genes

marker_genes_dict = {
'Oligodendrocytes':['PLP1','MOG','ST18','MBP',	'OLIG1',	'COL18A1','OPALIN','CNP'],
'OPC': ['PDGFRA','VCAN','CSPG4'],
'Astrocytes':['SLC1A2','GFAP','CD44','CHI3L1','AQP4']
'Microglia':['CSF1R','CX3CR1','TREM2','TYROBP','P2RY12'],
'Endothelial': 	['FLT1',	'ERG'],
'pericytes':	['NOTCH3'],
'VLMC':	['CYP1B1',	'COL1A2'],
'Excitatory':	['SLC17A7','CUX2','FEZF2','THEMIS','RORB','PCP4'],
'Inhibitory':['GAD1','GAD2','ADARB2','LHX6','SST','PVALB','KIT',"VIP"]           
             }


### Annotate the major celltypes
sc.pl.dotplot(adata, marker_genes_dict,'leiden_1.0', dendrogram=True,standard_scale="var")
plt.savefig("path/marker_harmony_leiden_1.0_dotplot.png")

### plot the Assay, total_counts
sc.pl.umap(adata,color=["Assay", "leiden_1.0","total_counts"], frameon=False,)
plt.savefig("path/plot_new_leiden1.0_Assay_leiden.png")

#### calculate Marker genes
adata.uns['log1p']["base"] = None
sc.tl.rank_genes_groups(adata,"leiden_1.0",method='wilcoxon',pts=True,corr_method='benjamini-hochberg')
sc.tl.filter_rank_genes_groups(adata,min_in_group_fraction=0.2,min_fold_change=0.8)
sc.tl.dendrogram(adata,"leiden_1.0")
sc.pl.rank_genes_groups_dotplot(adata,n_genes=25,key='rank_genes_groups_filtered',cmap='bwr',standard_scale='var')

#### save the marker genes
result = adata.uns['rank_genes_groups_filtered']
groups = result['names'].dtype.names
ranked_genes_df = pd.DataFrame(
    {group + '_' + key: result[key][group]
     for group in groups for key in ['names', 'pvals_adj', 'logfoldchanges']}
)

workbook = Workbook()
list_of_dfs = []
for val in adata.obs["leiden_1.0"].values.unique():
    print(val)
    sheet = workbook.create_sheet(title=val)
    sheet.title = val
    df_temp = sc.get.rank_genes_groups_df(adata,val,key="rank_genes_groups_filtered")
    df_temp.index = df_temp["names"]
    df_temp = df_temp.set_axis(str(val) + "_" + df_temp.columns.values, axis=1)
    for row in dataframe_to_rows(df_temp, index=False, header=True):
            #print(row)
            sheet.append(row)
workbook.save("path/markers.xlsx")


# ---- cell 10 ----
import scanpy.external as sce
import anndata
import pandas as pd
import scanpy as sc
import numpy as np
from scipy.io import mmread
from sklearn.decomposition import TruncatedSVD
import matplotlib
import matplotlib.pyplot as plt
import seaborn as sns
import csv
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#### Read the integrated data
adata=anndata.read_h5ad("path/integrated_harmony_adata.h5ad")

##### subset the excitatory,inhibitory,astrocytes,microglia,oligodendrocytes and identify the subtypes

exc=adata[adata.obs["major_celltypes"].isin(["excitatory"])]
inh=adata[adata.obs["major_celltypes"].isin(["Inhibitory"])]
Astro=adata[adata.obs["major_celltypes"].isin(["Astrocytes"])]
microglia=adata[adata.obs["major_celltypes"].isin(["Microglia"])]
oligo=adata[adata.obs["major_celltypes"].isin(["Oligodendrocytes"])]

#### Do the same for inhibitory neuronal population change the anndata subset "exc" to "inh"
#### PCA 
sc.tl.pca(exc, svd_solver='arpack')
sc.pl.pca_variance_ratio(exc, log=True)

### Harmony integration mention covariates

sce.pp.harmony_integrate(exc, ['Assay','Brain.Region'])

#### check the Elbow plot for PCA
### generate leiden graph based on KNN

sc.pp.neighbors(exc, use_rep="X_pca_harmony",n_pcs=30)
sc.tl.leiden(exc,resolution = 1.0, key_added = "leiden_1.0")
sc.tl.umap(exc,min_dist=0.6,spread=1.4)
sc.pl.umap(exc,color=["Assay", "leiden_1.0"],ncols=2, frameon=False,)
### identify subcluster specific markers 
exc.uns['log1p']["base"] = None
sc.tl.rank_genes_groups(exc,"leiden_1.0",method='wilcoxon',pts=True,corr_method='benjamini-hochberg')
sc.tl.filter_rank_genes_groups(exc,min_in_group_fraction=0.2,min_fold_change=0.8)
sc.tl.dendrogram(exc,"leiden_1.0")

#### calculate marker and plot them
sc.pl.rank_genes_groups_dotplot(exc,n_genes=15,key='rank_genes_groups_filtered',cmap='bwr',standard_scale='var')

###### Subsets are determined through a process involving correlograms, Pearson correlations between clusters, 
###### and the identification of distinctive marker genes.

result = exc.uns['rank_genes_groups_filtered']
groups = result['names'].dtype.names
ranked_genes_df = pd.DataFrame(
    {group + '_' + key: result[key][group]
     for group in groups for key in ['names', 'pvals_adj', 'logfoldchanges']}
)
workbook = Workbook()
list_of_dfs = []
for val in exc.obs["leiden_1.0"].values.unique():
    print(val)
    sheet = workbook.create_sheet(title=val)
    sheet.title = val
    df_temp = sc.get.rank_genes_groups_df(exc,val,key="rank_genes_groups_filtered")
    df_temp.index = df_temp["names"]
    df_temp = df_temp.set_axis(str(val) + "_" + df_temp.columns.values, axis=1)
    for row in dataframe_to_rows(df_temp, index=False, header=True):
            #print(row)
            sheet.append(row)
workbook.save("path/exc_markers.xlsx")



# ---- cell 11 ----
#### specify the non neuronal subset and  do the below for subclustering

sc.tl.pca(Astro, svd_solver='arpack')
sc.pl.pca_variance_ratio(Astro, log=True)

### Harmony integration mention covariates

sce.pp.harmony_integrate(Astro, ['Assay','Brain.Region'])

#### check the Elbow plot for PCA
### generate leiden graph based nearsest neighbor for excitatory

sc.pp.neighbors(Astro, use_rep="X_pca_harmony",n_pcs=12)


#### change the resolution "Astrocyte-0.3, microglia-0.2, oligodendrocytes-0.2"
sc.tl.leiden(Astro,resolution = 0.3, key_added = "leiden_0.3")
sc.tl.umap(Astro,min_dist=0.6,spread=1.2)
sc.pl.umap(Astro,color=["Assay", "leiden_0.3"],ncols=2, frameon=False,)

### identify subcluster specific markers 
Astro.uns['log1p']["base"] = None
sc.tl.rank_genes_groups(Astro,"leiden_0.3",method='wilcoxon',pts=True,corr_method='benjamini-hochberg')
sc.tl.filter_rank_genes_groups(Astro,min_in_group_fraction=0.2,min_fold_change=0.5)
sc.tl.dendrogram(Astro,"leiden_0.3")

#### calculate marker and plot them
sc.pl.rank_genes_groups_dotplot(Astro,n_genes=15,key='rank_genes_groups_filtered',cmap='bwr',standard_scale='var')

#### Annotate the anndata based on  markers and include them in anndata . obs slot "Author_Annotation"
#### save the marker genes with pts save for each non neuronal cluster marker genes

result = Astro.uns['rank_genes_groups_filtered']
groups = result['names'].dtype.names
ranked_genes_df = pd.DataFrame(
    {group + '_' + key: result[key][group]
     for group in groups for key in ['names', 'pvals_adj', 'logfoldchanges']}
)
workbook = Workbook()
list_of_dfs = []
for val in Astro.obs["leiden_0.3"].values.unique():
    print(val)
    sheet = workbook.create_sheet(title=val)
    sheet.title = val
    df_temp = sc.get.rank_genes_groups_df(Astro,val,key="rank_genes_groups_filtered")
    df_temp.index = df_temp["names"]
    df_temp = df_temp.set_axis(str(val) + "_" + df_temp.columns.values, axis=1)
    for row in dataframe_to_rows(df_temp, index=False, header=True):
            #print(row)
            sheet.append(row)
workbook.save("path/Astro_markers.xlsx")


#### combine all the anndata objects neuronal and non-neuronal 



